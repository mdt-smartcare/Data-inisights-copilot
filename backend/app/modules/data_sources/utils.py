"""
Utilities for data source processing.

DuckDB file handling, schema normalization, and helpers for large file processing.
"""
import os
import re
import csv
import json
import shutil
import logging
import unicodedata
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import duckdb
import threading

# Global lock for DuckDB write operations to prevent concurrent access issues
_duckdb_write_lock = threading.Lock()

from app.core.config import get_settings

logger = logging.getLogger(__name__)

# ==========================================
# Schema Normalization
# ==========================================

KNOWN_ABBREVIATIONS = {
    'bmi': 'bmi', 'bp': 'bp', 'hr': 'hr', 'id': 'id',
    'dob': 'dob', 'ssn': 'ssn', 'mrn': 'mrn', 'icd': 'icd',
    'cpt': 'cpt', 'npi': 'npi', 'ehr': 'ehr', 'emr': 'emr',
    'hba1c': 'hba1c', 'ldl': 'ldl', 'hdl': 'hdl',
    'ast': 'ast', 'alt': 'alt', 'wbc': 'wbc', 'rbc': 'rbc',
}

MAX_COLUMN_LENGTH = 63


def normalize_column_name(col: str, index: int = 0) -> str:
    """
    Normalize a column name to a SQL-safe identifier.
    
    Transformations:
    1. Strip whitespace
    2. Convert to lowercase
    3. Replace spaces and special chars with underscores
    4. Remove parentheses content or convert to suffix
    5. Collapse multiple underscores
    6. Ensure doesn't start with number
    7. Truncate long names
    """
    if not col or not col.strip():
        return f"col_{index}"
    
    # Normalize unicode
    name = unicodedata.normalize('NFKD', col)
    name = name.encode('ASCII', 'ignore').decode('ASCII')
    
    # Lowercase and strip
    name = name.lower().strip()
    
    # Handle parentheses - convert to suffix
    name = re.sub(r'\(([^)]+)\)', r'_\1', name)
    
    # Replace non-alphanumeric with underscore
    name = re.sub(r'[^a-z0-9_]', '_', name)
    
    # Collapse multiple underscores
    name = re.sub(r'_+', '_', name)
    
    # Strip leading/trailing underscores
    name = name.strip('_')
    
    # Ensure doesn't start with number
    if name and name[0].isdigit():
        name = f"col_{name}"
    
    # Handle empty result
    if not name:
        name = f"col_{index}"
    
    # Truncate if too long
    if len(name) > MAX_COLUMN_LENGTH:
        name = name[:MAX_COLUMN_LENGTH]
    
    return name


def normalize_table_name(filename: str) -> str:
    """Convert filename to valid SQL table name."""
    # Remove extension
    name = os.path.splitext(filename)[0]
    return normalize_column_name(name)


# ==========================================
# Path Helpers
# ==========================================

def get_agent_data_dir(agent_id: str) -> Path:
    """Get the directory for an agent's data files."""
    settings = get_settings()
    agent_dir = settings.duckdb_path / f"agent_{agent_id}"
    agent_dir.mkdir(parents=True, exist_ok=True)
    return agent_dir


def get_agent_duckdb_path(agent_id: str) -> Path:
    """Get path to an agent's DuckDB file."""
    return get_agent_data_dir(agent_id) / "database.duckdb"


def get_agent_csv_path(agent_id: str, table_name: str) -> Path:
    """Get path where an agent's CSV file will be stored."""
    return get_agent_data_dir(agent_id) / f"{table_name}.csv"


# ==========================================
# DataSource-based Path Helpers (primary)
# ==========================================

def get_datasource_dir(data_source_id: str) -> Path:
    """
    Get directory for a data source's files.
    
    Structure: data/duckdb_files/ds_{data_source_id}/
    """
    settings = get_settings()
    ds_dir = settings.duckdb_path / f"ds_{data_source_id}"
    ds_dir.mkdir(parents=True, exist_ok=True)
    return ds_dir


def get_datasource_duckdb_path(data_source_id: str) -> Path:
    """Get path to a data source's DuckDB file."""
    return get_datasource_dir(data_source_id) / "database.duckdb"


def get_datasource_csv_path(data_source_id: str) -> Path:
    """Get path where a data source's CSV file will be stored."""
    return get_datasource_dir(data_source_id) / "data.csv"


def get_datasource_source_path(data_source_id: str, original_filename: str) -> Path:
    """Get path where the original uploaded file will be stored."""
    return get_datasource_dir(data_source_id) / f"_source_{original_filename}"


def get_relative_datasource_duckdb_path(data_source_id: str) -> str:
    """
    Get RELATIVE path for storing in database.
    
    Returns path relative to settings.duckdb_path, e.g.:
    'ds_bd042320-7412-4a18-8395-e808fc24a18a/database.duckdb'
    """
    return f"ds_{data_source_id}/database.duckdb"


# Legacy user-based functions (deprecated - for backward compatibility)
def get_user_data_dir(user_id: str) -> Path:
    """DEPRECATED: Use get_datasource_dir instead."""
    settings = get_settings()
    user_dir = settings.duckdb_path / f"user_{user_id}"
    user_dir.mkdir(parents=True, exist_ok=True)
    return user_dir


def get_user_duckdb_path(user_id: str) -> Path:
    """DEPRECATED: Use get_datasource_duckdb_path instead."""
    return get_user_data_dir(user_id) / "database.duckdb"


# ==========================================
# Relative Path Helpers (for portable storage)
# ==========================================

def resolve_duckdb_path(relative_or_absolute_path: str) -> Path:
    """
    Resolve a duckdb_file_path to its full absolute path.
    
    Handles:
    - DataSource paths (new): 'ds_xxx/database.duckdb'
    - User paths (legacy): 'user_xxx/database.duckdb'
    - Absolute paths (legacy): 'D:\\...\\user_xxx\\database.duckdb'
    
    Returns the resolved absolute Path.
    """
    from pathlib import Path
    
    path = Path(relative_or_absolute_path)
    
    # If it's already absolute and exists, use it directly
    if path.is_absolute():
        if path.exists():
            return path
        # Absolute path doesn't exist - try to extract relative portion and resolve
        # Look for 'ds_', 'user_' or 'agent_' pattern in the path
        path_str = str(path)
        for marker in ['ds_', 'user_', 'agent_']:
            if marker in path_str:
                # Extract from the marker onwards
                idx = path_str.find(marker)
                relative = path_str[idx:]
                # Normalize path separators
                relative = relative.replace('\\', '/')
                settings = get_settings()
                resolved = settings.duckdb_path / relative
                if resolved.exists():
                    return resolved
        # Fallback: return the original path (will fail with file not found downstream)
        return path
    
    # Relative path - resolve from settings.duckdb_path
    settings = get_settings()
    return settings.duckdb_path / relative_or_absolute_path


# ==========================================
# Excel to CSV Streaming
# ==========================================

def stream_excel_to_csv(
    xlsx_path: str,
    csv_path: str,
    chunk_log_interval: int = 100000,
    estimated_rows: int = None,
    progress_tracker: Dict[str, Any] = None,
) -> Dict[str, Any]:
    """
    Stream Excel file to CSV using openpyxl read-only mode.
    Avoids loading entire Excel file into RAM.
    
    Args:
        xlsx_path: Path to Excel file
        csv_path: Output CSV path
        chunk_log_interval: Log progress every N rows
        estimated_rows: Estimated total rows for progress calculation
        progress_tracker: Optional shared dict for thread-safe progress updates.
                         Keys: 'rows_processed', 'total_rows', 'done'
    
    Returns:
        Dict with columns, row_count, elapsed_seconds
    """
    from openpyxl import load_workbook
    
    logger.info(f"Starting Excel → CSV streaming: {xlsx_path}")
    start_time = datetime.now()
    
    # Initialize progress tracker if provided
    if progress_tracker is not None:
        progress_tracker['rows_processed'] = 0
        progress_tracker['total_rows'] = estimated_rows or 0
        progress_tracker['done'] = False
    
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    
    headers: List[str] = []
    row_count = 0
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [
                    normalize_column_name(str(cell) if cell else f"col_{i}", i)
                    for i, cell in enumerate(row)
                ]
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
            else:
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = '' if cell is None else str(cell)
                writer.writerow(row_dict)
                row_count += 1
                
                # Update shared progress tracker (thread-safe for simple dict updates)
                if progress_tracker is not None:
                    progress_tracker['rows_processed'] = row_count
                
                if row_count % chunk_log_interval == 0:
                    elapsed = (datetime.now() - start_time).total_seconds()
                    rate = row_count / elapsed if elapsed > 0 else 0
                    progress_pct = f" ({row_count * 100 // estimated_rows}%)" if estimated_rows else ""
                    logger.info(f"  Processed {row_count:,} rows{progress_pct} ({rate:,.0f} rows/sec)")
    
    wb.close()
    
    # Mark as done
    if progress_tracker is not None:
        progress_tracker['rows_processed'] = row_count
        progress_tracker['done'] = True
    
    elapsed = (datetime.now() - start_time).total_seconds()
    logger.info(f"Excel → CSV complete: {row_count:,} rows in {elapsed:.1f}s")
    
    return {
        "columns": headers,
        "row_count": row_count,
        "elapsed_seconds": elapsed,
    }


# ==========================================
# DuckDB Operations
# ==========================================

def register_csv_in_duckdb(
    data_source_id: str,
    csv_path: str,
    original_filename: str,
    columns: List[str],
    row_count: int,
) -> None:
    """
    Register a CSV file in DuckDB as a virtual table.
    DuckDB queries CSV directly from disk without loading into RAM.
    
    Each data source gets its own DuckDB file with a single 'data' table.
    
    Uses a lock to prevent concurrent connection conflicts.
    """
    db_path = get_datasource_duckdb_path(data_source_id)
    
    # Use lock to prevent "different configuration" errors when multiple
    # connections try to access the same database file
    with _duckdb_write_lock:
        conn = duckdb.connect(str(db_path), read_only=False)
        
        try:
            # Create metadata table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS _file_metadata (
                    original_filename VARCHAR,
                    file_type VARCHAR,
                    csv_path VARCHAR,
                    row_count BIGINT,
                    columns JSON,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Clear old metadata and view
            conn.execute("DELETE FROM _file_metadata")
            conn.execute("DROP VIEW IF EXISTS data")
            
            # Create VIEW that reads directly from CSV (virtualized)
            csv_path_escaped = str(csv_path).replace("'", "''")
            conn.execute(f"""
                CREATE VIEW data AS 
                SELECT * FROM read_csv_auto('{csv_path_escaped}', header=true)
            """)
            
            # Store metadata
            conn.execute("""
                INSERT INTO _file_metadata (original_filename, file_type, csv_path, row_count, columns)
                VALUES (?, ?, ?, ?, ?)
            """, [original_filename, 'csv', str(csv_path), row_count, json.dumps(columns)])
            
            logger.info(f"Registered CSV as DuckDB view for data_source {data_source_id} ({row_count:,} rows)")
        
        finally:
            conn.close()


def process_file_for_duckdb_sync(
    data_source_id: str,
    source_path: str,
    file_type: str,
    original_filename: str,
    estimated_rows: int = 0,
) -> Dict[str, Any]:
    """
    Process a file (CSV/Excel) and register in DuckDB.
    Purely synchronous - safe to run in threading.Thread.
    
    Each data source gets its own folder with:
    - database.duckdb (DuckDB with VIEW 'data')
    - data.csv (the CSV file DuckDB reads)
    - _source_* (original uploaded file)
    """
    import time
    import psycopg2
    from concurrent.futures import ThreadPoolExecutor
    
    logger.info(f"Background processing started: {original_filename}")
    
    # Progress phases allocation
    PHASE_CONVERSION = 80  # Excel/CSV processing: 0-80%
    PHASE_REGISTRATION = 20  # DuckDB registration: 80-100%
    
    last_reported_progress = [0]
    last_reported_status = [None]
    
    def update_status(status: str, progress: int = None, error: str = None, row_count: int = None):
        """Update processing status synchronously using psycopg2."""
        # Always update on status change or terminal states
        status_changed = status != last_reported_status[0]
        is_terminal = status in ("completed", "failed")
        
        # Only skip if same status AND progress hasn't changed by at least 5%
        if not status_changed and not is_terminal:
            if progress is not None and abs(progress - last_reported_progress[0]) < 5 and progress < 100:
                return
        
        last_reported_progress[0] = progress or 0
        last_reported_status[0] = status
        
        logger.info(f"Updating status: data_source={data_source_id}, status={status}, progress={progress}")
        
        try:
            settings = get_settings()
            db_url = settings.postgres_uri
            
            conn = psycopg2.connect(db_url)
            conn.autocommit = True
            cur = conn.cursor()
            
            try:
                if status == "completed":
                    cur.execute("""
                        UPDATE data_sources 
                        SET processing_status = %s,
                            processing_progress = 100,
                            processing_error = NULL,
                            row_count = COALESCE(%s, row_count)
                        WHERE id = %s::uuid
                    """, (status, row_count, data_source_id))
                elif status == "failed":
                    cur.execute("""
                        UPDATE data_sources 
                        SET processing_status = %s,
                            processing_error = %s
                        WHERE id = %s::uuid
                    """, (status, error, data_source_id))
                else:
                    cur.execute("""
                        UPDATE data_sources 
                        SET processing_status = %s,
                            processing_progress = COALESCE(%s, processing_progress)
                        WHERE id = %s::uuid
                    """, (status, progress, data_source_id))
            finally:
                cur.close()
                conn.close()
        except Exception as e:
            logger.error(f"Failed to update status: {e}", exc_info=True)
    
    try:
        csv_path = get_datasource_csv_path(data_source_id)
        
        if file_type == 'xlsx':
            logger.info(f"Starting Excel processing for: {original_filename}")
            update_status("processing", progress=0)
            
            # Progress tracker for thread-safe updates
            progress_tracker: Dict[str, Any] = {
                'rows_processed': 0,
                'total_rows': estimated_rows,
                'done': False,
            }
            
            # Run Excel conversion in a thread pool so we can poll progress
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(
                    stream_excel_to_csv,
                    source_path,
                    str(csv_path),
                    100000,  # chunk_log_interval
                    estimated_rows,
                    progress_tracker,
                )
                
                # Poll progress while conversion runs
                while not progress_tracker['done']:
                    time.sleep(0.5)  # Check every 500ms
                    
                    if progress_tracker['total_rows'] > 0:
                        actual_progress = int(
                            (progress_tracker['rows_processed'] / progress_tracker['total_rows']) 
                            * PHASE_CONVERSION
                        )
                        update_status("processing", progress=actual_progress)
                
                # Get result (raises if there was an exception)
                result = future.result()
                columns = result["columns"]
                row_count = result["row_count"]
            
            update_status("processing", progress=PHASE_CONVERSION)
            
        elif file_type == 'csv':
            logger.info(f"Processing CSV (instant): {original_filename}")
            update_status("processing", progress=0)
            
            csv_path = Path(source_path)  # Use source file directly
            
            # Get columns
            conn = duckdb.connect(":memory:")
            csv_path_escaped = str(csv_path).replace("'", "''")
            
            if estimated_rows > 0:
                row_count = estimated_rows
            else:
                row_count = conn.execute(
                    f"SELECT COUNT(*) FROM read_csv_auto('{csv_path_escaped}', header=true)"
                ).fetchone()[0]
            
            cols = conn.execute(
                f"DESCRIBE SELECT * FROM read_csv_auto('{csv_path_escaped}', header=true, sample_size=1000)"
            ).fetchall()
            columns = [normalize_column_name(c[0], i) for i, c in enumerate(cols)]
            conn.close()
            
            update_status("processing", progress=PHASE_CONVERSION)
        else:
            raise ValueError(f"Unsupported file type for DuckDB: {file_type}")
        
        # DuckDB registration
        if file_type == 'xlsx':
            update_status("processing", progress=PHASE_CONVERSION + (PHASE_REGISTRATION // 2))
        
        register_csv_in_duckdb(
            data_source_id=data_source_id,
            csv_path=str(csv_path),
            original_filename=original_filename,
            columns=columns,
            row_count=row_count,
        )
        
        update_status("completed", row_count=row_count)
        logger.info(f"Background processing completed: {original_filename}")
        
        return {
            "columns": columns,
            "row_count": row_count,
            "csv_path": str(csv_path),
            "duckdb_path": get_relative_datasource_duckdb_path(data_source_id),
        }
    
    except Exception as e:
        logger.error(f"Error processing file for DuckDB: {e}", exc_info=True)
        update_status("failed", error=str(e))
        raise


# Keep the async version for potential future use, but it's not used currently
async def process_file_for_duckdb(
    data_source_id: str,
    source_path: str,
    file_type: str,
    original_filename: str,
    estimated_rows: int = 0,
) -> Dict[str, Any]:
    """
    Async version - delegates to sync version via to_thread.
    """
    import asyncio
    return await asyncio.to_thread(
        process_file_for_duckdb_sync,
        data_source_id,
        source_path,
        file_type,
        original_filename,
        estimated_rows,
    )


def execute_duckdb_query(
    data_source_id: str,
    query: str,
    max_rows: int = 10000,
) -> Dict[str, Any]:
    """
    Execute a SQL query against a data source's DuckDB.
    
    Note: The table is always named 'data' in each data source's DuckDB.
    
    Returns:
        Dict with status, columns, rows, row_count, execution_time_ms
    """
    import time
    
    db_path = get_datasource_duckdb_path(data_source_id)
    
    if not db_path.exists():
        return {
            "status": "error",
            "error": "Data source not found or not yet processed.",
            "columns": [],
            "rows": [],
            "row_count": 0,
        }
    
    try:
        start_time = time.time()
        
        conn = duckdb.connect(str(db_path), read_only=True)
        result = conn.execute(query)
        
        columns = [desc[0] for desc in result.description]
        rows_data = result.fetchmany(max_rows)
        has_more = len(rows_data) == max_rows
        
        conn.close()
        
        execution_time_ms = (time.time() - start_time) * 1000
        
        # Convert to list of dicts with JSON-safe values
        rows = []
        for row in rows_data:
            row_dict = {}
            for i, col in enumerate(columns):
                val = row[i]
                if val is None:
                    row_dict[col] = None
                elif isinstance(val, (int, float, str, bool)):
                    row_dict[col] = val
                else:
                    row_dict[col] = str(val)
            rows.append(row_dict)
        
        return {
            "status": "success",
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "execution_time_ms": round(execution_time_ms, 2),
            "error": f"Results limited to {max_rows} rows" if has_more else None,
        }
        
    except duckdb.Error as e:
        return {
            "status": "error",
            "error": f"SQL Error: {str(e)}",
            "columns": [],
            "rows": [],
            "row_count": 0,
        }
    except Exception as e:
        return {
            "status": "error",
            "error": f"Query failed: {str(e)}",
            "columns": [],
            "rows": [],
            "row_count": 0,
        }


def get_datasource_metadata(data_source_id: str) -> Optional[Dict[str, Any]]:
    """Get metadata for a data source's DuckDB."""
    db_path = get_datasource_duckdb_path(data_source_id)
    
    if not db_path.exists():
        return None
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        tables_exist = conn.execute("""
            SELECT COUNT(*) FROM information_schema.tables 
            WHERE table_name = '_file_metadata'
        """).fetchone()[0]
        
        if not tables_exist:
            conn.close()
            return None
        
        row = conn.execute("""
            SELECT original_filename, file_type, row_count, columns, created_at
            FROM _file_metadata
            LIMIT 1
        """).fetchone()
        
        conn.close()
        
        if not row:
            return None
            
        return {
            "original_filename": row[0],
            "file_type": row[1],
            "row_count": row[2],
            "columns": json.loads(row[3]) if row[3] else [],
            "created_at": str(row[4]) if row[4] else None,
        }
        
    except Exception as e:
        logger.error(f"Failed to get metadata for data_source {data_source_id}: {e}")
        return None


def get_datasource_schema(data_source_id: str) -> Optional[List[Dict[str, str]]]:
    """Get schema (columns and types) for a data source."""
    db_path = get_datasource_duckdb_path(data_source_id)
    
    if not db_path.exists():
        return None
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        result = conn.execute("DESCRIBE SELECT * FROM data").fetchall()
        conn.close()
        
        return [
            {"column_name": row[0], "data_type": row[1]}
            for row in result
        ]
    except Exception as e:
        logger.error(f"Failed to get schema for data_source {data_source_id}: {e}")
        return None


def delete_datasource_files(data_source_id: str) -> bool:
    """Delete all files for a data source (DuckDB, CSV, original file)."""
    ds_dir = get_datasource_dir(data_source_id)
    
    # Don't create the directory if it doesn't exist
    settings = get_settings()
    actual_dir = settings.duckdb_path / f"ds_{data_source_id}"
    
    if not actual_dir.exists():
        return True
    
    try:
        shutil.rmtree(actual_dir, ignore_errors=True)
        logger.info(f"Deleted all files for data_source {data_source_id}")
        return True
    except Exception as e:
        logger.error(f"Failed to delete data_source files: {e}")
        return False


# Legacy functions (deprecated - for backward compatibility)
def list_duckdb_tables(user_id: str) -> List[Dict[str, Any]]:
    """DEPRECATED: Use get_datasource_metadata instead."""
    db_path = get_user_duckdb_path(user_id)
    
    if not db_path.exists():
        return []
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        tables_exist = conn.execute("""
            SELECT COUNT(*) FROM information_schema.tables 
            WHERE table_name = '_file_metadata'
        """).fetchone()[0]
        
        if not tables_exist:
            conn.close()
            return []
        
        rows = conn.execute("""
            SELECT table_name, original_filename, file_type, row_count, columns, created_at
            FROM _file_metadata
            ORDER BY created_at DESC
        """).fetchall()
        
        conn.close()
        
        return [
            {
                "name": row[0],
                "original_filename": row[1],
                "file_type": row[2],
                "row_count": row[3],
                "columns": json.loads(row[4]) if row[4] else [],
                "created_at": str(row[5]) if row[5] else None,
            }
            for row in rows
        ]
        
    except Exception as e:
        logger.error(f"Failed to list tables: {e}")
        return []


def get_table_schema(user_id: str, table_name: str) -> Optional[List[Dict[str, str]]]:
    """DEPRECATED: Use get_datasource_schema instead."""
    db_path = get_user_duckdb_path(user_id)
    
    if not db_path.exists():
        return None
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        result = conn.execute(f"DESCRIBE SELECT * FROM {table_name}").fetchall()
        conn.close()
        
        return [
            {"column_name": row[0], "data_type": row[1]}
            for row in result
        ]
    except Exception as e:
        logger.error(f"Failed to get schema for {table_name}: {e}")
        return None


def delete_duckdb_table(user_id: str, table_name: str) -> bool:
    """DEPRECATED: Use delete_datasource_files instead."""
    db_path = get_user_duckdb_path(user_id)
    
    if not db_path.exists():
        return False
    
    try:
        with _duckdb_write_lock:
            conn = duckdb.connect(str(db_path), read_only=False)
        
        csv_info = conn.execute(
            "SELECT csv_path FROM _file_metadata WHERE table_name = ?",
            [table_name]
        ).fetchone()
        
        if not csv_info:
            conn.close()
            return False
        
        csv_path = csv_info[0]
        
        conn.execute(f"DROP VIEW IF EXISTS {table_name}")
        conn.execute("DELETE FROM _file_metadata WHERE table_name = ?", [table_name])
        conn.close()
        
        if csv_path and os.path.exists(csv_path):
            os.remove(csv_path)
        
        logger.info(f"Table deleted: {table_name}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to delete table: {e}")
        return False


def delete_all_user_tables(user_id: str) -> bool:
    """DEPRECATED: Use delete_datasource_files instead."""
    user_dir = get_user_data_dir(user_id)
    
    if not user_dir.exists():
        return True
    
    try:
        shutil.rmtree(user_dir, ignore_errors=True)
        logger.info(f"All tables deleted for user {user_id}")
        return True
    except Exception as e:
        logger.error(f"Failed to delete user data: {e}")
        return False


# ==========================================
# Fast Column Extraction (for large files)
# ==========================================

def extract_file_metadata_fast(file_path: str, file_type: str) -> Dict[str, Any]:
    """
    Extract column names, types, and row count from a file in a single pass.
    
    For CSV: Uses DuckDB to get columns and count in one connection.
    For Excel: Uses openpyxl read-only mode to read header + max_row metadata.
    
    Args:
        file_path: Path to the file
        file_type: File type ('csv' or 'xlsx')
        
    Returns:
        Dict with:
        - columns: List[str] - normalized column names
        - column_details: List[Dict[str, str]] - dicts with 'name' and 'type' keys
        - row_count: int - number of data rows (excluding header)
    """
    if file_type == 'csv':
        return _extract_csv_metadata_fast(file_path)
    elif file_type == 'xlsx':
        return _extract_excel_metadata_fast(file_path)
    else:
        return {"columns": [], "column_details": [], "row_count": 0}


def _extract_csv_metadata_fast(file_path: str) -> Dict[str, Any]:
    """Extract columns and row count from CSV using DuckDB in one connection."""
    try:
        conn = duckdb.connect(":memory:")
        csv_path_escaped = str(file_path).replace("'", "''")
        
        # Get columns with DESCRIBE (uses sample_size internally)
        cols_result = conn.execute(
            f"DESCRIBE SELECT * FROM read_csv_auto('{csv_path_escaped}', header=true, sample_size=1000)"
        ).fetchall()
        
        # Get exact row count
        row_count = conn.execute(
            f"SELECT COUNT(*) FROM read_csv_auto('{csv_path_escaped}', header=true)"
        ).fetchone()[0]
        
        conn.close()
        
        columns = []
        column_details = []
        
        for i, row in enumerate(cols_result):
            original_name = row[0]
            normalized_name = normalize_column_name(original_name, i)
            col_type = str(row[1]).upper() if row[1] else 'VARCHAR'
            
            columns.append(normalized_name)
            column_details.append({"name": normalized_name, "type": col_type})
        
        logger.info(f"CSV metadata: {len(columns)} columns, {row_count:,} rows - {file_path}")
        return {"columns": columns, "column_details": column_details, "row_count": row_count}
        
    except Exception as e:
        logger.error(f"Failed to extract CSV metadata: {e}")
        return {"columns": [], "column_details": [], "row_count": 0}


def _extract_excel_metadata_fast(file_path: str) -> Dict[str, Any]:
    """Extract columns and row count from Excel in one workbook load."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        columns = []
        column_details = []
        
        # Get header row
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, cell in enumerate(row):
                original_name = str(cell) if cell else f"col_{i}"
                normalized_name = normalize_column_name(original_name, i)
                
                columns.append(normalized_name)
                column_details.append({"name": normalized_name, "type": "VARCHAR"})
            break
        
        # Get row count from metadata (instant, no iteration needed)
        row_count = ws.max_row - 1 if ws.max_row else 0  # Subtract header row
        
        wb.close()
        
        logger.info(f"Excel metadata: {len(columns)} columns, {row_count:,} rows - {file_path}")
        return {"columns": columns, "column_details": column_details, "row_count": max(row_count, 0)}
        
    except Exception as e:
        logger.error(f"Failed to extract Excel metadata: {e}")
        return {"columns": [], "column_details": [], "row_count": 0}


# Legacy wrappers (for backward compatibility)
def extract_file_columns_fast(file_path: str, file_type: str) -> tuple:
    """
    DEPRECATED: Use extract_file_metadata_fast instead.
    
    Returns:
        Tuple of (column_names: List[str], column_details: List[Dict[str, str]])
    """
    result = extract_file_metadata_fast(file_path, file_type)
    return result["columns"], result["column_details"]


# ==========================================
# Database URL Security
# ==========================================

import base64


def mask_db_url(db_url: str) -> str:
    """
    Mask credentials in a database URL.
    
    Example:
        postgresql://user:password@localhost:5433/db
        → postgresql://***:***@localhost:5433/db
    """
    return re.sub(r'://[^:]+:[^@]+@', '://***:***@', db_url)


def decode_db_url(encoded_url: str) -> str:
    """
    Decode a base64 encoded database URL received from frontend.
    
    Raises:
        ValueError: If decoding fails
    """
    try:
        return base64.b64decode(encoded_url.encode('utf-8')).decode('utf-8')
    except Exception as e:
        raise ValueError(f"Invalid encoded database URL: {e}")
