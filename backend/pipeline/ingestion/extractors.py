"""
Concrete extractors for the Multi-Modal Data Ingestion Engine.

Each extractor implements :class:`BaseExtractor` and yields a stream of
:class:`Document` objects from a specific file format.  All extractors use
``yield`` to remain memory-efficient and wrap processing in ``try/except``
blocks so that corrupted or unreadable chunks are logged and skipped rather
than crashing the pipeline.
"""

from __future__ import annotations

import csv
import json
import logging
import os
from typing import Any, Generator

from backend.pipeline.ingestion.models import BaseExtractor, Document

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# PDF Extractor
# ---------------------------------------------------------------------------

class PDFExtractor(BaseExtractor):
    """Extracts one :class:`Document` per page from a PDF file.

    Uses `pdfplumber` for robust text extraction including tables and
    complex layouts.
    """

    def extract(self, file_path: str) -> Generator[Document, None, None]:
        """Yield a Document for every page in *file_path*.

        Args:
            file_path: Path to a ``.pdf`` file.

        Yields:
            Document with ``page_content`` set to the page text and metadata
            containing ``source``, ``page_number``, and ``total_pages``.
        """
        import pdfplumber

        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"PDF file not found: {file_path}")

        try:
            with pdfplumber.open(file_path) as pdf:
                total_pages = len(pdf.pages)
                for page_num, page in enumerate(pdf.pages, start=1):
                    try:
                        text = page.extract_text() or ""
                        text = text.strip()
                        if not text:
                            logger.warning(
                                "Page %d of '%s' produced no text — skipping.",
                                page_num,
                                file_path,
                            )
                            continue
                        yield Document(
                            page_content=text,
                            metadata={
                                "source": file_path,
                                "file_type": "pdf",
                                "page_number": page_num,
                                "total_pages": total_pages,
                            },
                        )
                    except Exception as exc:
                        logger.error(
                            "Error extracting page %d from '%s': %s",
                            page_num,
                            file_path,
                            exc,
                        )
        except Exception as exc:
            logger.error("Failed to open PDF '%s': %s", file_path, exc)
            raise ValueError(f"Cannot parse PDF file: {file_path}") from exc


# ---------------------------------------------------------------------------
# CSV Extractor
# ---------------------------------------------------------------------------

class CSVExtractor(BaseExtractor):
    """Extracts one :class:`Document` per row from a CSV file.

    Uses the stdlib ``csv.DictReader`` to stream rows without loading the
    entire file into memory.  Each row is formatted as human-readable
    ``key: value`` lines.
    """

    def __init__(self, encoding: str = "utf-8-sig") -> None:
        self.encoding = encoding

    def extract(self, file_path: str) -> Generator[Document, None, None]:
        """Yield a Document for every row in *file_path*.

        Args:
            file_path: Path to a ``.csv`` file.

        Yields:
            Document with row content and metadata containing ``source``,
            ``row_number``, and ``columns``.
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"CSV file not found: {file_path}")

        try:
            with open(file_path, "r", encoding=self.encoding, newline="") as fh:
                reader = csv.DictReader(fh)
                columns = reader.fieldnames or []
                for row_num, row in enumerate(reader, start=1):
                    try:
                        content_parts = [
                            f"{key}: {value}"
                            for key, value in row.items()
                            if value and str(value).strip()
                        ]
                        if not content_parts:
                            continue
                        yield Document(
                            page_content="\n".join(content_parts),
                            metadata={
                                "source": file_path,
                                "file_type": "csv",
                                "row_number": row_num,
                                "columns": list(columns),
                            },
                        )
                    except Exception as exc:
                        logger.error(
                            "Error processing row %d in '%s': %s",
                            row_num,
                            file_path,
                            exc,
                        )
        except UnicodeDecodeError as exc:
            logger.error(
                "Encoding error reading '%s' with encoding '%s': %s",
                file_path,
                self.encoding,
                exc,
            )
            raise ValueError(
                f"Cannot decode CSV file '{file_path}' with encoding "
                f"'{self.encoding}'"
            ) from exc
        except Exception as exc:
            logger.error("Failed to open CSV '%s': %s", file_path, exc)
            raise ValueError(f"Cannot parse CSV file: {file_path}") from exc


# ---------------------------------------------------------------------------
# Excel Extractor
# ---------------------------------------------------------------------------

class ExcelExtractor(BaseExtractor):
    """Extracts one :class:`Document` per row per sheet from an Excel file.

    Uses ``openpyxl`` in **read-only** mode to stream rows without loading
    the full workbook into memory.
    """

    def extract(self, file_path: str) -> Generator[Document, None, None]:
        """Yield a Document for every row in every sheet of *file_path*.

        Args:
            file_path: Path to a ``.xlsx`` file.

        Yields:
            Document with row content and metadata containing ``source``,
            ``sheet_name``, ``row_number``, and ``columns``.
        """
        from openpyxl import load_workbook

        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("Failed to open Excel file '%s': %s", file_path, exc)
            raise ValueError(f"Cannot parse Excel file: {file_path}") from exc

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers: list[str] = []
                for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    try:
                        if row_num == 1:
                            headers = [
                                str(cell) if cell is not None else f"col_{i}"
                                for i, cell in enumerate(row)
                            ]
                            continue

                        content_parts = []
                        for header, cell_value in zip(headers, row):
                            if cell_value is not None and str(cell_value).strip():
                                content_parts.append(f"{header}: {cell_value}")

                        if not content_parts:
                            continue

                        yield Document(
                            page_content="\n".join(content_parts),
                            metadata={
                                "source": file_path,
                                "file_type": "xlsx",
                                "sheet_name": sheet_name,
                                "row_number": row_num,
                                "columns": list(headers),
                            },
                        )
                    except Exception as exc:
                        logger.error(
                            "Error processing row %d in sheet '%s' of '%s': %s",
                            row_num,
                            sheet_name,
                            file_path,
                            exc,
                        )
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# JSON Extractor
# ---------------------------------------------------------------------------

class JSONExtractor(BaseExtractor):
    """Extracts :class:`Document` objects from a JSON file.

    Handles two common shapes:

    * **Array of objects** — yields one Document per object.
    * **Single object** — yields one Document for the entire file.

    Nested keys are flattened into dot-notation (e.g. ``address.city``).
    """

    def extract(self, file_path: str) -> Generator[Document, None, None]:
        """Yield Documents from *file_path*.

        Args:
            file_path: Path to a ``.json`` file.

        Yields:
            Document with flattened key-value content.
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"JSON file not found: {file_path}")

        try:
            with open(file_path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
        except json.JSONDecodeError as exc:
            logger.error("Invalid JSON in '%s': %s", file_path, exc)
            raise ValueError(f"Cannot parse JSON file: {file_path}") from exc
        except Exception as exc:
            logger.error("Failed to open JSON file '%s': %s", file_path, exc)
            raise ValueError(f"Cannot read JSON file: {file_path}") from exc

        if isinstance(data, list):
            for idx, item in enumerate(data):
                try:
                    if not isinstance(item, dict):
                        item = {"value": item}
                    flat = self._flatten(item)
                    content = "\n".join(f"{k}: {v}" for k, v in flat.items() if v)
                    if not content:
                        continue
                    yield Document(
                        page_content=content,
                        metadata={
                            "source": file_path,
                            "file_type": "json",
                            "item_index": idx,
                        },
                    )
                except Exception as exc:
                    logger.error(
                        "Error processing item %d in '%s': %s",
                        idx,
                        file_path,
                        exc,
                    )
        elif isinstance(data, dict):
            flat = self._flatten(data)
            content = "\n".join(f"{k}: {v}" for k, v in flat.items() if v)
            if content:
                yield Document(
                    page_content=content,
                    metadata={
                        "source": file_path,
                        "file_type": "json",
                    },
                )
        else:
            logger.warning(
                "JSON root in '%s' is neither a dict nor a list — skipping.",
                file_path,
            )

    # -- helpers -------------------------------------------------------------

    @staticmethod
    def _flatten(
        obj: dict[str, Any],
        parent_key: str = "",
        sep: str = ".",
    ) -> dict[str, str]:
        """Recursively flatten a nested dict into dot-notation keys."""
        items: list[tuple[str, str]] = []
        for key, value in obj.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else key
            if isinstance(value, dict):
                items.extend(
                    JSONExtractor._flatten(value, new_key, sep).items()
                )
            elif isinstance(value, list):
                items.append((new_key, ", ".join(str(v) for v in value)))
            else:
                items.append((new_key, str(value) if value is not None else ""))
        return dict(items)
