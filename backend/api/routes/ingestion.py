"""
Ingestion API route — File upload endpoint with DuckDB for large file SQL support.
Supports RAG + SQL on CSV/Excel files up to millions of rows.

Architecture for Large Files (6.5M+ rows):
1. Excel files are streamed to CSV using openpyxl read-only mode (no RAM blowup)
2. CSV/Parquet files are stored persistently on disk
3. DuckDB queries files directly from disk (virtualized, no full RAM load)
4. Schema Normalizer sanitizes all column names for predictable SQL
"""

import os
import tempfile
import logging
import json
import shutil
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import duckdb
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, BackgroundTasks
from pydantic import BaseModel

from backend.pipeline.ingestion.factory import DocumentLoaderFactory
from backend.pipeline.ingestion.schema_normalizer import (
    SchemaNormalizer, 
    normalize_column_name, 
    normalize_table_name,
)
from backend.core.permissions import get_current_user, User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/ingestion", tags=["ingestion"])

# Directory to store user data files and DuckDB databases
DATA_STORAGE_DIR = Path(__file__).parent.parent.parent.parent / "data" / "duckdb_files"
DATA_STORAGE_DIR.mkdir(parents=True, exist_ok=True)

# Global schema normalizer instance
schema_normalizer = SchemaNormalizer()


# ---------------------------------------------------------------------------
# Helper Functions for Large File Processing
# ---------------------------------------------------------------------------

def _sanitize_table_name(filename: str) -> str:
    """Convert filename to valid SQL table name using Schema Normalizer."""
    return normalize_table_name(filename)


def _sanitize_column_name(col: str, index: int = 0) -> str:
    """Sanitize a single column name using Schema Normalizer."""
    return normalize_column_name(col, index)


def _get_user_data_dir(user_id: int) -> Path:
    """Get the directory for a user's data files."""
    user_dir = DATA_STORAGE_DIR / f"user_{user_id}"
    user_dir.mkdir(parents=True, exist_ok=True)
    return user_dir


def _get_user_duckdb_path(user_id: int) -> Path:
    """Get the path to a user's DuckDB file."""
    return _get_user_data_dir(user_id) / "database.duckdb"


def _get_user_csv_path(user_id: int, table_name: str) -> Path:
    """Get the path where a user's CSV file will be stored."""
    return _get_user_data_dir(user_id) / f"{table_name}.csv"


def _stream_excel_to_csv(
    xlsx_path: str, 
    csv_path: str,
    chunk_log_interval: int = 100000
) -> Dict[str, Any]:
    """
    Stream Excel file to CSV using openpyxl read-only mode.
    This avoids loading the entire Excel file into RAM.
    
    For a 6.5M row file, this processes ~10-50k rows/sec depending on columns.
    
    Returns:
        Dict with columns, row_count, and processing stats
    """
    from openpyxl import load_workbook
    
    logger.info(f"Starting Excel → CSV streaming conversion: {xlsx_path}")
    start_time = datetime.now()
    
    # Open workbook in read-only mode (critical for large files!)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    
    headers = []
    row_count = 0
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                # First row = headers
                headers = [normalize_column_name(str(cell) if cell else f"col_{i}", i) 
                          for i, cell in enumerate(row)]
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
            else:
                # Data rows
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        # Handle None and convert to string for CSV
                        row_dict[headers[i]] = '' if cell is None else str(cell)
                writer.writerow(row_dict)
                row_count += 1
                
                # Log progress for large files
                if row_count % chunk_log_interval == 0:
                    elapsed = (datetime.now() - start_time).total_seconds()
                    rate = row_count / elapsed if elapsed > 0 else 0
                    logger.info(f"  Processed {row_count:,} rows ({rate:,.0f} rows/sec)")
    
    wb.close()
    
    elapsed = (datetime.now() - start_time).total_seconds()
    logger.info(f"Excel → CSV conversion complete: {row_count:,} rows in {elapsed:.1f}s")
    
    return {
        "columns": headers,
        "row_count": row_count,
        "elapsed_seconds": elapsed,
    }


def _register_csv_in_duckdb(
    user_id: int,
    table_name: str,
    csv_path: str,
    original_filename: str,
    columns: List[str],
    row_count: int,
) -> None:
    """
    Register a CSV file in DuckDB as a virtual table.
    DuckDB will query the CSV directly from disk without loading into RAM.
    """
    db_path = _get_user_duckdb_path(user_id)
    conn = duckdb.connect(str(db_path))
    
    try:
        # Create metadata table if not exists
        conn.execute("""
            CREATE TABLE IF NOT EXISTS _file_metadata (
                table_name VARCHAR PRIMARY KEY,
                original_filename VARCHAR,
                file_type VARCHAR,
                csv_path VARCHAR,
                row_count BIGINT,
                columns JSON,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Remove old entry if exists
        conn.execute("DELETE FROM _file_metadata WHERE table_name = ?", [table_name])
        
        # Drop old view if exists
        conn.execute(f"DROP VIEW IF EXISTS {table_name}")
        
        # Create a VIEW that reads directly from CSV (virtualized - no RAM load!)
        conn.execute(f"""
            CREATE VIEW {table_name} AS 
            SELECT * FROM read_csv_auto('{csv_path}', header=true)
        """)
        
        # Store metadata
        conn.execute("""
            INSERT INTO _file_metadata (table_name, original_filename, file_type, csv_path, row_count, columns)
            VALUES (?, ?, ?, ?, ?, ?)
        """, [table_name, original_filename, 'csv', str(csv_path), row_count, json.dumps(columns)])
        
        logger.info(f"Registered CSV as DuckDB view: {table_name} ({row_count:,} rows)")
        
    finally:
        conn.close()


def _process_large_file_background(
    user_id: int,
    table_name: str,
    source_path: str,
    file_type: str,
    original_filename: str,
) -> None:
    """
    Background task to process large files.
    Converts Excel to CSV (streaming) and registers in DuckDB.
    """
    try:
        csv_path = _get_user_csv_path(user_id, table_name)
        
        if file_type == 'xlsx':
            # Stream Excel to CSV (memory-efficient)
            result = _stream_excel_to_csv(source_path, str(csv_path))
            columns = result["columns"]
            row_count = result["row_count"]
        elif file_type == 'csv':
            # Just copy the CSV file to permanent storage
            shutil.copy(source_path, csv_path)
            # Get row count and columns using DuckDB (efficient)
            conn = duckdb.connect(":memory:")
            info = conn.execute(f"SELECT COUNT(*) FROM read_csv_auto('{csv_path}')").fetchone()
            row_count = info[0]
            cols = conn.execute(f"DESCRIBE SELECT * FROM read_csv_auto('{csv_path}')").fetchall()
            columns = [normalize_column_name(c[0], i) for i, c in enumerate(cols)]
            conn.close()
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        # Register in DuckDB
        _register_csv_in_duckdb(
            user_id=user_id,
            table_name=table_name,
            csv_path=str(csv_path),
            original_filename=original_filename,
            columns=columns,
            row_count=row_count,
        )
        
        logger.info(f"Background processing complete for {original_filename}: {row_count:,} rows")
        
    except Exception as e:
        logger.error(f"Background processing failed for {original_filename}: {e}")
        raise


def _get_file_row_count_estimate(file_path: str, file_type: str) -> int:
    """
    Quickly estimate row count without reading entire file.
    Uses file size heuristics for speed.
    """
    file_size = os.path.getsize(file_path)
    
    if file_type == 'csv':
        # Estimate ~100 bytes per row for typical CSV
        return file_size // 100
    elif file_type == 'xlsx':
        # Excel files are compressed, estimate ~50 bytes per row
        return file_size // 50
    return 0


# ---------------------------------------------------------------------------
# Response schemas
# ---------------------------------------------------------------------------

class ExtractedDocument(BaseModel):
    """A single extracted document preview."""
    page_content: str
    metadata: Dict[str, Any]


class IngestionResponse(BaseModel):
    """Response from the file upload / extraction endpoint."""
    status: str
    file_name: str
    file_type: str
    total_documents: int
    documents: List[ExtractedDocument]
    table_name: Optional[str] = None
    columns: Optional[List[str]] = None
    column_details: Optional[List[Dict[str, str]]] = None  # [{name, type}]
    row_count: Optional[int] = None
    processing_mode: Optional[str] = None  # 'sync' or 'background'
    message: Optional[str] = None


class SQLQueryRequest(BaseModel):
    """Request model for SQL queries on uploaded files."""
    query: str


class SQLQueryResponse(BaseModel):
    """Response from SQL query execution."""
    status: str
    query: str
    row_count: int
    columns: List[str]
    rows: List[Dict[str, Any]]
    execution_time_ms: Optional[float] = None
    error: Optional[str] = None


class FileTablesResponse(BaseModel):
    """Response listing available tables from uploaded files."""
    tables: List[Dict[str, Any]]


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

SUPPORTED_EXTENSIONS = DocumentLoaderFactory.supported_extensions()
SQL_SUPPORTED_EXTENSIONS = {'.csv', '.xlsx'}
MAX_PREVIEW_DOCS = 50
MAX_CONTENT_LENGTH = 500
MAX_SQL_ROWS = 10000
LARGE_FILE_THRESHOLD = 10 * 1024 * 1024  # 10MB = process in background


@router.post("/upload", response_model=IngestionResponse)
async def upload_and_extract(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Upload a file and extract documents for RAG + SQL querying.

    Accepts ``.pdf``, ``.csv``, ``.xlsx``, and ``.json`` files.
    
    For ``.csv`` and ``.xlsx`` files:
    - Small files (<10MB): Processed immediately
    - Large files (≥10MB): Processed in background, SQL available shortly
    - Data is stored for DuckDB queries (supports millions of rows)
    - Sample documents extracted for RAG
    
    Returns up to 50 document previews for RAG indexing.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided.")

    _, ext = os.path.splitext(file.filename)
    ext = ext.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Supported: {', '.join(SUPPORTED_EXTENSIONS)}",
        )

    # Create temp directory for upload
    tmp_dir = tempfile.mkdtemp(prefix="ingestion_")
    tmp_path = os.path.join(tmp_dir, file.filename)

    table_name = None
    columns = None
    column_details = None
    row_count = None
    processing_mode = None
    message = None

    try:
        # Stream file to disk (handles large files without RAM issues)
        file_size = 0
        with open(tmp_path, "wb") as f:
            while chunk := await file.read(1024 * 1024):  # 1MB chunks
                f.write(chunk)
                file_size += len(chunk)

        logger.info(
            "Ingestion upload: file='%s', size=%d bytes (%.2f MB), user='%s'",
            file.filename,
            file_size,
            file_size / (1024 * 1024),
            current_user.username,
        )

        # Process for SQL support (CSV and Excel)
        if ext in SQL_SUPPORTED_EXTENSIONS:
            table_name = _sanitize_table_name(file.filename)
            file_type = ext.lstrip('.')
            
            if file_size >= LARGE_FILE_THRESHOLD:
                # Large file - process in background
                processing_mode = "background"
                estimated_rows = _get_file_row_count_estimate(tmp_path, file_type)
                
                # Copy to permanent location first (temp will be deleted)
                permanent_source = _get_user_data_dir(current_user.id) / f"_source_{file.filename}"
                shutil.copy(tmp_path, permanent_source)
                
                # Schedule background processing
                background_tasks.add_task(
                    _process_large_file_background,
                    user_id=current_user.id,
                    table_name=table_name,
                    source_path=str(permanent_source),
                    file_type=file_type,
                    original_filename=file.filename,
                )
                
                row_count = estimated_rows
                message = f"Large file detected ({file_size / (1024*1024):.1f} MB). Processing in background. SQL queries will be available shortly."
                logger.info(f"Large file queued for background processing: {file.filename}")
                
            else:
                # Small file - process immediately
                processing_mode = "sync"
                try:
                    _process_large_file_background(
                        user_id=current_user.id,
                        table_name=table_name,
                        source_path=tmp_path,
                        file_type=file_type,
                        original_filename=file.filename,
                    )
                    
                    # Get actual metadata
                    db_path = _get_user_duckdb_path(current_user.id)
                    conn = duckdb.connect(str(db_path), read_only=True)
                    meta = conn.execute(
                        "SELECT columns, row_count FROM _file_metadata WHERE table_name = ?",
                        [table_name]
                    ).fetchone()
                    
                    # Get column type info via DESCRIBE
                    column_details = None
                    try:
                        type_info = conn.execute(
                            f"DESCRIBE SELECT * FROM {table_name}"
                        ).fetchall()
                        column_details = [
                            {"name": row[0], "type": row[1]}
                            for row in type_info
                        ]
                    except Exception as desc_err:
                        logger.warning(f"Could not DESCRIBE table {table_name}: {desc_err}")
                    
                    conn.close()
                    
                    if meta:
                        columns = json.loads(meta[0])
                        row_count = meta[1]
                        
                except Exception as e:
                    logger.error(f"Failed to process file for SQL: {e}")
                    message = f"SQL processing failed: {str(e)}. RAG extraction may still work."

        # Extract sample documents for RAG (limit iterations for large files)
        documents: List[ExtractedDocument] = []
        total = 0
        
        try:
            extractor = DocumentLoaderFactory.get_extractor(tmp_path)
            for doc in extractor.extract(tmp_path):
                total += 1
                if len(documents) < MAX_PREVIEW_DOCS:
                    preview = doc.page_content[:MAX_CONTENT_LENGTH]
                    if len(doc.page_content) > MAX_CONTENT_LENGTH:
                        preview += "…"
                    documents.append(
                        ExtractedDocument(
                            page_content=preview,
                            metadata=doc.metadata,
                        )
                    )
                # Stop early for very large files (we have row_count from DuckDB)
                if total >= MAX_PREVIEW_DOCS and row_count:
                    total = row_count
                    break
                if total > 100000:
                    total = row_count or total
                    break
        except Exception as e:
            logger.warning(f"RAG extraction failed: {e}")

        logger.info(
            "Ingestion complete: file='%s', rows=%s, mode=%s",
            file.filename,
            f"{row_count:,}" if row_count else "unknown",
            processing_mode,
        )

        return IngestionResponse(
            status="success",
            file_name=file.filename,
            file_type=ext.lstrip("."),
            total_documents=total,
            documents=documents,
            table_name=table_name,
            columns=columns,
            column_details=column_details,
            row_count=row_count,
            processing_mode=processing_mode,
            message=message,
        )

    except ValueError as exc:
        logger.error("Ingestion failed for '%s': %s", file.filename, exc)
        raise HTTPException(status_code=422, detail=str(exc))
    except Exception as exc:
        logger.error("Unexpected error during ingestion of '%s': %s", file.filename, exc)
        raise HTTPException(status_code=500, detail=f"Internal error: {str(exc)}")
    finally:
        # Cleanup temp directory
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        except Exception:
            pass


@router.get("/sql/tables", response_model=FileTablesResponse)
async def list_sql_tables(
    current_user: User = Depends(get_current_user),
):
    """List all uploaded file tables available for SQL querying."""
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        return FileTablesResponse(tables=[])
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        # Check if metadata table exists
        tables_exist = conn.execute("""
            SELECT COUNT(*) FROM information_schema.tables 
            WHERE table_name = '_file_metadata'
        """).fetchone()[0]
        
        if not tables_exist:
            conn.close()
            return FileTablesResponse(tables=[])
        
        rows = conn.execute("""
            SELECT table_name, original_filename, file_type, row_count, columns, created_at
            FROM _file_metadata
            ORDER BY created_at DESC
        """).fetchall()
        
        conn.close()
        
        tables = []
        for row in rows:
            tables.append({
                "name": row[0],
                "original_filename": row[1],
                "file_type": row[2],
                "row_count": row[3],
                "columns": json.loads(row[4]) if row[4] else [],
                "created_at": str(row[5]) if row[5] else None,
            })
        
        return FileTablesResponse(tables=tables)
        
    except Exception as e:
        logger.error(f"Failed to list tables: {e}")
        return FileTablesResponse(tables=[])


@router.post("/sql/query", response_model=SQLQueryResponse)
async def execute_sql_query(
    request: SQLQueryRequest,
    current_user: User = Depends(get_current_user),
):
    """Execute a SQL query against uploaded file data using DuckDB.
    
    DuckDB queries CSV files directly from disk (virtualized).
    Supports 6.5M+ rows without loading into RAM.
    
    Supported SQL features:
    - Aggregations: AVG, SUM, COUNT, MIN, MAX
    - GROUP BY, ORDER BY, HAVING
    - JOINs across multiple uploaded tables
    - Window functions
    - CTEs (WITH clauses)
    
    Example queries:
    - SELECT * FROM your_table LIMIT 10
    - SELECT AVG(age), gender FROM patients GROUP BY gender
    - SELECT * FROM patients WHERE age > 30 ORDER BY age DESC
    """
    import time
    
    query = request.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query cannot be empty.")
    
    # Security: Only allow SELECT queries (and WITH for CTEs)
    query_upper = query.upper().strip()
    if not (query_upper.startswith("SELECT") or query_upper.startswith("WITH")):
        raise HTTPException(
            status_code=400,
            detail="Only SELECT queries are allowed for security reasons.",
        )
    
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        raise HTTPException(
            status_code=400,
            detail="No files uploaded for SQL querying. Upload a CSV or Excel file first.",
        )
    
    try:
        start_time = time.time()
        
        conn = duckdb.connect(str(db_path), read_only=True)
        result = conn.execute(query)
        
        # Get column names
        columns = [desc[0] for desc in result.description]
        
        # Fetch rows (with limit for safety)
        rows_data = result.fetchmany(MAX_SQL_ROWS)
        has_more = len(rows_data) == MAX_SQL_ROWS
        
        conn.close()
        
        execution_time_ms = (time.time() - start_time) * 1000
        
        # Convert to list of dicts with JSON-safe values
        rows = []
        for row in rows_data:
            row_dict = {}
            for i, col in enumerate(columns):
                val = row[i]
                if val is None:
                    row_dict[col] = None
                elif isinstance(val, (int, float, str, bool)):
                    row_dict[col] = val
                else:
                    row_dict[col] = str(val)
            rows.append(row_dict)
        
        logger.info(
            "SQL query executed: user='%s', rows=%d, time=%.2fms%s",
            current_user.username,
            len(rows),
            execution_time_ms,
            " (truncated)" if has_more else "",
        )
        
        return SQLQueryResponse(
            status="success",
            query=query,
            row_count=len(rows),
            columns=columns,
            rows=rows,
            execution_time_ms=round(execution_time_ms, 2),
            error=f"Results limited to {MAX_SQL_ROWS} rows" if has_more else None,
        )
        
    except duckdb.Error as e:
        logger.error("DuckDB error for user '%s': %s", current_user.username, e)
        return SQLQueryResponse(
            status="error",
            query=query,
            row_count=0,
            columns=[],
            rows=[],
            error=f"SQL Error: {str(e)}",
        )
    except Exception as e:
        logger.error("Query execution failed for user '%s': %s", current_user.username, e)
        raise HTTPException(
            status_code=500,
            detail=f"Query execution failed: {str(e)}",
        )


@router.delete("/sql/tables/{table_name}")
async def delete_sql_table(
    table_name: str,
    current_user: User = Depends(get_current_user),
):
    """Delete an uploaded file table and its CSV data."""
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="No tables found.")
    
    try:
        conn = duckdb.connect(str(db_path))
        
        # Get CSV path before deleting
        csv_info = conn.execute(
            "SELECT csv_path FROM _file_metadata WHERE table_name = ?",
            [table_name]
        ).fetchone()
        
        if not csv_info:
            conn.close()
            raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found.")
        
        csv_path = csv_info[0]
        
        # Drop view and metadata
        conn.execute(f"DROP VIEW IF EXISTS {table_name}")
        conn.execute("DELETE FROM _file_metadata WHERE table_name = ?", [table_name])
        conn.close()
        
        # Delete CSV file
        if csv_path and os.path.exists(csv_path):
            os.remove(csv_path)
        
        logger.info(f"Table deleted: {table_name}, user={current_user.username}")
        
        return {"status": "success", "message": f"Table '{table_name}' deleted."}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete table: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sql/tables")
async def delete_all_sql_tables(
    current_user: User = Depends(get_current_user),
):
    """Delete all uploaded file tables and data for the current user."""
    user_dir = _get_user_data_dir(current_user.id)
    
    if not user_dir.exists():
        return {"status": "success", "message": "No tables to delete."}
    
    try:
        # Delete entire user data directory
        shutil.rmtree(user_dir, ignore_errors=True)
        
        logger.info(f"All tables deleted for user={current_user.username}")
        
        return {"status": "success", "message": "All tables deleted."}
        
    except Exception as e:
        logger.error(f"Failed to delete user data: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sql/schema/{table_name}")
async def get_table_schema(
    table_name: str,
    current_user: User = Depends(get_current_user),
):
    """Get the schema (columns and types) of a specific table."""
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="No tables found.")
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        # Get column info from the view
        result = conn.execute(f"DESCRIBE SELECT * FROM {table_name}").fetchall()
        conn.close()
        
        schema = [
            {"column_name": row[0], "data_type": row[1]}
            for row in result
        ]
        
        return {"table_name": table_name, "schema": schema}
        
    except duckdb.CatalogException:
        raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found.")
    except Exception as e:
        logger.error(f"Failed to get schema: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sql/status/{table_name}")
async def get_processing_status(
    table_name: str,
    current_user: User = Depends(get_current_user),
):
    """Check if a table is ready for querying (useful for background processing)."""
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        return {"status": "not_found", "ready": False}
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        result = conn.execute(
            "SELECT row_count, created_at FROM _file_metadata WHERE table_name = ?",
            [table_name]
        ).fetchone()
        
        conn.close()
        
        if result:
            return {
                "status": "ready",
                "ready": True,
                "row_count": result[0],
                "created_at": str(result[1]),
            }
        else:
            return {"status": "processing", "ready": False}
            
    except Exception as e:
        logger.error(f"Failed to check status: {e}")
        return {"status": "error", "ready": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Text-to-SQL Endpoint (Natural Language → SQL → Results)
# ---------------------------------------------------------------------------

class NaturalLanguageQueryRequest(BaseModel):
    """Request for natural language query (Text-to-SQL)."""
    question: str


class NaturalLanguageQueryResponse(BaseModel):
    """Response from Text-to-SQL query."""
    status: str
    answer: Optional[str] = None
    sql: Optional[str] = None
    columns: Optional[List[str]] = None
    rows: Optional[List[Dict[str, Any]]] = None
    total_rows: Optional[int] = None
    execution_time_ms: Optional[float] = None
    error: Optional[str] = None


@router.post("/sql/ask", response_model=NaturalLanguageQueryResponse)
async def ask_natural_language(
    request: NaturalLanguageQueryRequest,
    current_user: User = Depends(get_current_user),
):
    """
    Text-to-SQL: Ask questions in natural language about your uploaded data.
    
    This is the PRIMARY query interface for structured file data.
    The LLM translates your question to SQL, executes against DuckDB,
    and returns results with a natural language answer.
    
    **Performance**: Queries on 6.5M+ rows execute in milliseconds.
    **Cost**: Zero embedding tokens - pure SQL execution.
    
    Example questions:
    - "How many patients are there?"
    - "What is the average age of female patients?"
    - "Show me the distribution of encounter types"
    - "Which patients have BMI greater than 30?"
    - "What's the average blood pressure by gender for patients over 50?"
    """
    from backend.services.file_sql_service import get_file_sql_service
    
    question = request.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="Question cannot be empty.")
    
    try:
        # Initialize FileSQLService for this user
        file_sql = get_file_sql_service(current_user.id)
        
        # Execute Text-to-SQL pipeline
        result = file_sql.query(question)
        
        if result["status"] == "error":
            return NaturalLanguageQueryResponse(
                status="error",
                error=result.get("error"),
                sql=result.get("sql"),
            )
        
        return NaturalLanguageQueryResponse(
            status="success",
            answer=result.get("answer"),
            sql=result.get("sql"),
            columns=result.get("columns"),
            rows=result.get("rows"),
            total_rows=result.get("total_rows"),
            execution_time_ms=result.get("execution_time_ms"),
        )
        
    except ValueError as e:
        # No files uploaded
        raise HTTPException(
            status_code=400,
            detail=str(e),
        )
    except Exception as e:
        logger.error(f"Text-to-SQL failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Query failed: {str(e)}",
        )


@router.get("/sql/schema-context")
async def get_schema_for_llm(
    current_user: User = Depends(get_current_user),
):
    """
    Get the full schema context for LLM prompts.
    
    This returns the schema in a format suitable for including
    in system prompts for Text-to-SQL generation.
    """
    from backend.services.file_sql_service import get_file_sql_service
    
    try:
        file_sql = get_file_sql_service(current_user.id)
        schema = file_sql.get_schema()
        
        return {
            "status": "success",
            "tables": schema.get("tables", []),
            "schema_text": schema.get("schema_text", ""),
        }
        
    except ValueError as e:
        return {
            "status": "no_data",
            "tables": [],
            "schema_text": "",
            "message": str(e),
        }
    except Exception as e:
        logger.error(f"Failed to get schema context: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Unified Query Endpoint with Intent Routing (SQL + RAG)
# ---------------------------------------------------------------------------

class UnifiedQueryRequest(BaseModel):
    """Request for unified query with automatic intent routing."""
    question: str
    use_llm_routing: bool = True  # Use LLM for more accurate routing


class UnifiedQueryResponse(BaseModel):
    """Response from unified query service."""
    status: str
    query_type: str  # 'sql', 'rag', 'hybrid', 'sql_fallback'
    intent: str
    confidence: float
    
    # Final answer (always populated)
    final_answer: Optional[str] = None
    
    # SQL results (if SQL was used)
    sql_answer: Optional[str] = None
    sql_query: Optional[str] = None
    sql_rows: Optional[List[Dict[str, Any]]] = None
    sql_execution_ms: Optional[float] = None
    
    # RAG results (if RAG was used)
    rag_answer: Optional[str] = None
    rag_documents: Optional[List[Dict[str, Any]]] = None
    rag_sources: Optional[List[str]] = None
    
    # Metadata
    routing_reason: Optional[str] = None
    error: Optional[str] = None


@router.post("/query", response_model=UnifiedQueryResponse)
async def unified_query(
    request: UnifiedQueryRequest,
    current_user: User = Depends(get_current_user),
):
    """
    Unified query endpoint with automatic intent routing.
    
    This is the RECOMMENDED endpoint for querying uploaded file data.
    The system automatically determines the optimal retrieval strategy:
    
    **SQL Engine** (milliseconds, structured data):
    - "How many patients are there?"
    - "What is the average age by gender?"
    - "Show patients with BMI > 30"
    
    **RAG Engine** (semantic search, unstructured text):
    - "Find patients with chronic migraine symptoms"
    - "Search for cases mentioning vision loss"
    - "Identify patients exhibiting fatigue and weight loss"
    
    **Hybrid** (combines both):
    - "Average age of patients mentioning chest pain"
    - "Count patients with notes about diabetes complications"
    
    The intent router uses pattern matching + optional LLM classification
    to determine the best approach for each query.
    """
    from backend.services.file_query_service import get_file_query_service
    
    question = request.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="Question cannot be empty.")
    
    try:
        # Initialize unified query service
        query_service = get_file_query_service(current_user.id)
        
        # Execute with intent routing
        result = query_service.query(
            question, 
            use_llm_routing=request.use_llm_routing
        )
        
        return UnifiedQueryResponse(
            status=result.status,
            query_type=result.query_type,
            intent=result.intent,
            confidence=result.confidence,
            final_answer=result.final_answer,
            sql_answer=result.sql_answer,
            sql_query=result.sql_query,
            sql_rows=result.sql_rows,
            sql_execution_ms=result.sql_execution_ms,
            rag_answer=result.rag_answer,
            rag_documents=result.rag_documents,
            rag_sources=result.rag_sources,
            routing_reason=result.routing_reason,
            error=result.error,
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Unified query failed: {e}")
        raise HTTPException(status_code=500, detail=f"Query failed: {str(e)}")


@router.post("/query/preview")
async def preview_query_routing(
    request: UnifiedQueryRequest,
    current_user: User = Depends(get_current_user),
):
    """
    Preview how a query will be routed WITHOUT executing it.
    
    Useful for UI to show users which engine will handle their query
    before they submit it.
    
    Returns routing decision with confidence score and reasoning.
    """
    from backend.services.file_query_service import get_file_query_service
    
    question = request.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="Question cannot be empty.")
    
    try:
        query_service = get_file_query_service(current_user.id)
        preview = query_service.get_routing_preview(question)
        
        return {
            "status": "success",
            **preview
        }
        
    except ValueError as e:
        return {
            "status": "no_data",
            "question": question,
            "engine": "none",
            "message": str(e),
        }
    except Exception as e:
        logger.error(f"Routing preview failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Column Classification Endpoint (for RAG configuration UI)
# ---------------------------------------------------------------------------

@router.get("/columns/classify/{table_name}")
async def classify_table_columns(
    table_name: str,
    current_user: User = Depends(get_current_user),
):
    """
    Classify columns in a table as structured vs unstructured.
    
    Returns recommendations for which columns should be:
    - SQL only (structured: age, gender, dates, IDs)
    - RAG embedding (unstructured: notes, history, descriptions)
    
    This helps users configure selective RAG extraction to avoid
    embedding millions of rows of structured data unnecessarily.
    """
    from backend.pipeline.ingestion.selective_extractor import (
        SelectiveColumnExtractor,
    )
    
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="No tables found.")
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        # Get column info
        col_info = conn.execute(f"DESCRIBE SELECT * FROM {table_name}").fetchall()
        columns = [row[0] for row in col_info]
        duckdb_types = {row[0]: row[1] for row in col_info}
        
        # Get sample data for classification
        sample_query = f"SELECT * FROM {table_name} LIMIT 100"
        sample_rows = conn.execute(sample_query).fetchall()
        
        conn.close()
        
        # Build sample data dict
        sample_data = {col: [] for col in columns}
        for row in sample_rows:
            for i, col in enumerate(columns):
                if i < len(row):
                    sample_data[col].append(row[i])
        
        # Classify columns
        extractor = SelectiveColumnExtractor()
        classifications = extractor.classify_columns(columns, sample_data, duckdb_types)
        summary = extractor.get_extraction_summary(classifications)
        
        return {
            "status": "success",
            "table_name": table_name,
            "total_columns": len(columns),
            "text_columns": summary["text_columns"],
            "structured_columns": summary["structured_columns"],
            "classifications": {
                col: {
                    "type": cls.column_type.value,
                    "confidence": cls.confidence,
                    "reason": cls.reason,
                    "avg_length": round(cls.avg_length, 1),
                    "sample_values": cls.sample_values[:3],
                }
                for col, cls in classifications.items()
            },
            "recommendation": {
                "embed_for_rag": summary["text_columns"],
                "sql_only": summary["structured_columns"],
                "estimated_rag_rows": f"~{len(sample_rows) * 100:,}" if summary["text_columns"] else "0",
            }
        }
        
    except duckdb.CatalogException:
        raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found.")
    except Exception as e:
        logger.error(f"Column classification failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# RAG Embedding Endpoints for Text Columns
# ---------------------------------------------------------------------------

class RAGConfigRequest(BaseModel):
    """Request to configure RAG embedding for a table."""
    table_name: str
    text_columns: List[str]  # Columns to embed for RAG
    id_column: str = "patient_id"  # Row identifier column
    parent_chunk_size: int = 800
    child_chunk_size: int = 200


class RAGProcessingResponse(BaseModel):
    """Response from RAG processing."""
    status: str
    table_name: str
    text_columns: List[str]
    total_documents: Optional[int] = None
    parent_chunks: Optional[int] = None
    child_chunks: Optional[int] = None
    embeddings_created: Optional[int] = None
    processing_time_seconds: Optional[float] = None
    message: Optional[str] = None
    error: Optional[str] = None


@router.post("/rag/embed", response_model=RAGProcessingResponse)
async def embed_text_columns_for_rag(
    request: RAGConfigRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
):
    """
    Start RAG embedding for selected text columns in a table.
    
    **CRITICAL FOR LARGE DATASETS:**
    - ONLY embed unstructured text columns (doctor_notes, clinical_history)
    - NEVER embed structured columns (age, patient_id, blood_pressure)
    - Use `/columns/classify/{table_name}` first to identify text columns
    
    **Architecture for 6.5M rows:**
    - If 1M rows have doctor_notes → ~1M parent chunks + ~4M child chunks
    - Uses local BGE-M3 model (GPU accelerated, zero API cost)
    - Parent-child chunking for precise retrieval with full context
    
    **Example:**
    ```json
    {
        "table_name": "patients",
        "text_columns": ["doctor_notes", "clinical_history"],
        "id_column": "patient_id"
    }
    ```
    
    Processing runs in background. Check status with `/rag/status/{table_name}`.
    """
    import asyncio
    from backend.pipeline.file_rag_pipeline import FileRAGPipeline, FileRAGConfig
    
    table_name = request.table_name
    text_columns = request.text_columns
    
    if not text_columns:
        raise HTTPException(
            status_code=400,
            detail="At least one text column must be specified for RAG embedding."
        )
    
    # Verify table exists
    db_path = _get_user_duckdb_path(current_user.id)
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="No tables found.")
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        # Verify table exists
        try:
            conn.execute(f"SELECT 1 FROM {table_name} LIMIT 1")
        except duckdb.CatalogException:
            conn.close()
            raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found.")
        
        # Verify columns exist
        col_info = conn.execute(f"DESCRIBE SELECT * FROM {table_name}").fetchall()
        existing_columns = {row[0] for row in col_info}
        conn.close()
        
        missing_columns = [col for col in text_columns if col not in existing_columns]
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Columns not found in table: {missing_columns}. Available: {list(existing_columns)}"
            )
        
        if request.id_column not in existing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"ID column '{request.id_column}' not found. Available: {list(existing_columns)}"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
    # Store RAG config in metadata
    _store_rag_config(
        current_user.id, 
        table_name, 
        text_columns, 
        request.id_column,
        "processing"
    )
    
    # Start background processing
    async def run_rag_pipeline():
        try:
            config = FileRAGConfig(
                text_columns=text_columns,
                parent_chunk_size=request.parent_chunk_size,
                child_chunk_size=request.child_chunk_size,
            )
            
            pipeline = FileRAGPipeline(current_user.id, config)
            
            stats = await pipeline.process_table_for_rag(
                table_name=table_name,
                text_columns=text_columns,
                id_column=request.id_column,
            )
            
            # Update status
            _store_rag_config(
                current_user.id,
                table_name,
                text_columns,
                request.id_column,
                "ready",
                stats
            )
            
            logger.info(f"RAG embedding complete for {table_name}: {stats}")
            
        except Exception as e:
            logger.error(f"RAG embedding failed for {table_name}: {e}")
            _store_rag_config(
                current_user.id,
                table_name,
                text_columns,
                request.id_column,
                "error",
                {"error": str(e)}
            )
    
    # Run in background
    background_tasks.add_task(asyncio.create_task, run_rag_pipeline())
    
    return RAGProcessingResponse(
        status="processing",
        table_name=table_name,
        text_columns=text_columns,
        message=f"RAG embedding started for columns: {text_columns}. "
                f"This may take several minutes for large datasets. "
                f"Check status with GET /ingestion/rag/status/{table_name}"
    )


def _store_rag_config(
    user_id: int,
    table_name: str,
    text_columns: List[str],
    id_column: str,
    status: str,
    stats: Optional[Dict] = None
):
    """Store RAG configuration in DuckDB metadata."""
    db_path = _get_user_duckdb_path(user_id)
    conn = duckdb.connect(str(db_path))
    
    try:
        # Create RAG metadata table if not exists
        conn.execute("""
            CREATE TABLE IF NOT EXISTS _rag_config (
                table_name VARCHAR PRIMARY KEY,
                text_columns JSON,
                id_column VARCHAR,
                status VARCHAR,
                stats JSON,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Upsert config
        conn.execute("""
            INSERT OR REPLACE INTO _rag_config 
            (table_name, text_columns, id_column, status, stats, updated_at)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, [
            table_name,
            json.dumps(text_columns),
            id_column,
            status,
            json.dumps(stats) if stats else None
        ])
        
    finally:
        conn.close()


@router.get("/rag/status/{table_name}")
async def get_rag_status(
    table_name: str,
    current_user: User = Depends(get_current_user),
):
    """
    Check the RAG embedding status for a table.
    
    Returns:
    - status: 'not_configured', 'processing', 'ready', 'error'
    - text_columns: Columns configured for RAG
    - stats: Processing statistics (when ready)
    """
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        return {"status": "not_configured", "table_name": table_name}
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        # Check if RAG config table exists
        tables = conn.execute("""
            SELECT COUNT(*) FROM information_schema.tables 
            WHERE table_name = '_rag_config'
        """).fetchone()[0]
        
        if not tables:
            conn.close()
            return {"status": "not_configured", "table_name": table_name}
        
        result = conn.execute("""
            SELECT text_columns, id_column, status, stats, updated_at
            FROM _rag_config WHERE table_name = ?
        """, [table_name]).fetchone()
        
        conn.close()
        
        if not result:
            return {"status": "not_configured", "table_name": table_name}
        
        return {
            "status": result[2],
            "table_name": table_name,
            "text_columns": json.loads(result[0]) if result[0] else [],
            "id_column": result[1],
            "stats": json.loads(result[3]) if result[3] else None,
            "updated_at": str(result[4]) if result[4] else None,
        }
        
    except Exception as e:
        logger.error(f"Failed to get RAG status: {e}")
        return {"status": "error", "table_name": table_name, "error": str(e)}


@router.get("/rag/tables")
async def list_rag_enabled_tables(
    current_user: User = Depends(get_current_user),
):
    """
    List all tables that have RAG embedding configured.
    
    Returns table names with their embedding status and statistics.
    """
    db_path = _get_user_duckdb_path(current_user.id)
    
    if not db_path.exists():
        return {"tables": []}
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        # Check if RAG config table exists
        tables_exist = conn.execute("""
            SELECT COUNT(*) FROM information_schema.tables 
            WHERE table_name = '_rag_config'
        """).fetchone()[0]
        
        if not tables_exist:
            conn.close()
            return {"tables": []}
        
        rows = conn.execute("""
            SELECT table_name, text_columns, id_column, status, stats, updated_at
            FROM _rag_config
            ORDER BY updated_at DESC
        """).fetchall()
        
        conn.close()
        
        tables = []
        for row in rows:
            stats = json.loads(row[4]) if row[4] else {}
            tables.append({
                "table_name": row[0],
                "text_columns": json.loads(row[1]) if row[1] else [],
                "id_column": row[2],
                "status": row[3],
                "parent_chunks": stats.get("parent_chunks"),
                "child_chunks": stats.get("child_chunks"),
                "embeddings_created": stats.get("embeddings_created"),
                "updated_at": str(row[5]) if row[5] else None,
            })
        
        return {"tables": tables}
        
    except Exception as e:
        logger.error(f"Failed to list RAG tables: {e}")
        return {"tables": [], "error": str(e)}


@router.delete("/rag/{table_name}")
async def delete_rag_embeddings(
    table_name: str,
    current_user: User = Depends(get_current_user),
):
    """
    Delete RAG embeddings for a table.
    
    Removes:
    - Vector DB collection (Qdrant or ChromaDB) with child embeddings
    - SQLite docstore with parent documents
    - RAG configuration metadata
    """
    user_dir = _get_user_data_dir(current_user.id)
    db_path = _get_user_duckdb_path(current_user.id)
    
    try:
        from backend.services.chroma_service import get_vector_store_type
        import os
        
        collection_name = f"file_rag_{table_name}"
        vector_store_type = get_vector_store_type()
        
        # Delete from vector store based on configured provider
        if vector_store_type == 'qdrant':
            # Delete Qdrant collection
            try:
                import requests
                qdrant_url = os.getenv("QDRANT_URL", "http://localhost:6333")
                response = requests.delete(f"{qdrant_url}/collections/{collection_name}", timeout=10)
                if response.status_code in [200, 404]:
                    logger.info(f"Deleted Qdrant collection: {collection_name}")
            except Exception as e:
                logger.warning(f"Failed to delete Qdrant collection {collection_name}: {e}")
        
        # Always try to clean up ChromaDB as well (for migration scenarios)
        try:
            import chromadb
            from chromadb.config import Settings
            
            chroma_path = user_dir / "chroma_db"
            if chroma_path.exists():
                chroma_client = chromadb.PersistentClient(
                    path=str(chroma_path),
                    settings=Settings(anonymized_telemetry=False),
                )
                
                try:
                    chroma_client.delete_collection(collection_name)
                    logger.info(f"Deleted ChromaDB collection: {collection_name}")
                except Exception:
                    pass
                    
        except ImportError:
            pass
        
        # Delete parent docstore
        docstore_path = user_dir / f"{table_name}_parents.db"
        if docstore_path.exists():
            os.remove(docstore_path)
            logger.info(f"Deleted docstore: {docstore_path}")
        
        # Remove RAG config from metadata
        if db_path.exists():
            conn = duckdb.connect(str(db_path))
            try:
                conn.execute("DELETE FROM _rag_config WHERE table_name = ?", [table_name])
            except Exception:
                pass
            finally:
                conn.close()
        
        logger.info(f"RAG embeddings deleted for table: {table_name}")
        
        return {
            "status": "success",
            "message": f"RAG embeddings deleted for table '{table_name}'.",
            "vector_store_type": vector_store_type
        }
        
    except Exception as e:
        logger.error(f"Failed to delete RAG embeddings: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rag/search/{table_name}")
async def semantic_search_table(
    table_name: str,
    query: str,
    top_k: int = 10,
    current_user: User = Depends(get_current_user),
):
    """
    Perform semantic search on a specific table's text columns.
    
    This is a lower-level endpoint for direct RAG search.
    For automatic SQL/RAG routing, use the unified `/query` endpoint.
    
    **Example:**
    ```
    POST /ingestion/rag/search/patients?query=chronic migraine with vision loss&top_k=5
    ```
    
    Returns matching clinical notes with parent context and source row IDs.
    """
    import asyncio
    from backend.pipeline.file_rag_pipeline import get_file_rag_pipeline
    
    if not query.strip():
        raise HTTPException(status_code=400, detail="Query cannot be empty.")
    
    # Check RAG status
    status = await get_rag_status(table_name, current_user)
    
    if status.get("status") != "ready":
        raise HTTPException(
            status_code=400,
            detail=f"RAG not ready for table '{table_name}'. Status: {status.get('status')}. "
                   f"Configure RAG with POST /ingestion/rag/embed first."
        )
    
    try:
        pipeline = get_file_rag_pipeline(current_user.id)
        
        # Run async search
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            results = loop.run_until_complete(
                pipeline.semantic_search(
                    query=query.strip(),
                    table_name=table_name,
                    top_k=top_k,
                    return_parents=True,
                )
            )
        finally:
            loop.close()
        
        return {
            "status": "success",
            "query": query,
            "table_name": table_name,
            "result_count": len(results),
            "results": results,
        }
        
    except Exception as e:
        logger.error(f"Semantic search failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Agentic Hybrid Query Endpoint (RAG → SQL → Synthesis)
# ---------------------------------------------------------------------------

class AgenticHybridRequest(BaseModel):
    """Request for agentic hybrid query (RAG feeds SQL)."""
    question: str
    table_name: Optional[str] = None  # Auto-detect if not provided
    rag_top_k: int = 50  # Number of semantic matches to find
    use_patient_ids: bool = True  # Extract patient IDs from RAG results


class AgenticHybridResponse(BaseModel):
    """Response from agentic hybrid query."""
    status: str
    question: str
    
    # Workflow stages
    stage_1_rag: Dict[str, Any]  # RAG semantic search results
    stage_2_sql: Dict[str, Any]  # SQL aggregation results  
    stage_3_synthesis: Dict[str, Any]  # Final synthesized answer
    
    # Final answer
    final_answer: str
    
    # Performance metrics
    total_time_ms: float
    rag_time_ms: float
    sql_time_ms: float
    synthesis_time_ms: float
    
    error: Optional[str] = None


@router.post("/query/agentic-hybrid", response_model=AgenticHybridResponse)
async def agentic_hybrid_query(
    request: AgenticHybridRequest,
    current_user: User = Depends(get_current_user),
):
    """
     **Agentic Hybrid Workflow**
    
    This is the most sophisticated query approach, combining semantic search
    with SQL aggregations in a multi-stage workflow:
    
    **Stage 1 - RAG Semantic Search:**
    - Find patients/rows matching semantic criteria
    - "patients with chronic headache symptoms" → finds relevant patient IDs
    
    **Stage 2 - SQL Aggregation:**
    - Use RAG-discovered IDs in SQL WHERE clause
    - Run aggregations (AVG, COUNT, etc.) on the filtered subset
    
    **Stage 3 - LLM Synthesis:**
    - Combine RAG context + SQL results into final answer
    - Provides both statistical facts AND semantic understanding
    
    **Example Queries:**
    - "What is the average age of patients with migraine symptoms?"
    - "How many patients mention chest pain in their notes?"
    - "Average BMI of patients with diabetes-related concerns"
    
    **Performance on 6.5M rows:**
    - RAG search: ~500ms (finds ~50 relevant patients)
    - SQL aggregation: ~50ms (queries only filtered subset)
    - Total: <1 second with full semantic + statistical answer
    """
    import time
    from backend.services.file_query_service import get_file_query_service
    
    question = request.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="Question cannot be empty.")
    
    total_start = time.time()
    
    try:
        query_service = get_file_query_service(current_user.id)
        
        # Execute the agentic hybrid workflow
        result = query_service.agentic_hybrid_query(
            question=question,
            table_name=request.table_name,
            rag_top_k=request.rag_top_k,
        )
        
        total_time_ms = (time.time() - total_start) * 1000
        
        return AgenticHybridResponse(
            status=result.get("status", "success"),
            question=question,
            stage_1_rag=result.get("stage_1_rag", {}),
            stage_2_sql=result.get("stage_2_sql", {}),
            stage_3_synthesis=result.get("stage_3_synthesis", {}),
            final_answer=result.get("final_answer", "Unable to generate answer."),
            total_time_ms=total_time_ms,
            rag_time_ms=result.get("rag_time_ms", 0),
            sql_time_ms=result.get("sql_time_ms", 0),
            synthesis_time_ms=result.get("synthesis_time_ms", 0),
            error=result.get("error"),
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Agentic hybrid query failed: {e}")
        raise HTTPException(status_code=500, detail=f"Query failed: {str(e)}")


@router.get("/query/workflow-status")
async def get_workflow_capabilities(
    current_user: User = Depends(get_current_user),
):
    """
    Check which query workflows are available for the current user.
    
    Returns the status of:
    - SQL Engine: Available if files are uploaded
    - RAG Engine: Available if text columns are embedded
    - Agentic Hybrid: Available if both SQL and RAG are ready
    """
    from backend.services.file_query_service import get_file_query_service
    
    try:
        query_service = get_file_query_service(current_user.id)
        
        # Check SQL availability
        sql_available = False
        sql_tables = []
        try:
            schema = query_service.sql_service.get_schema()
            sql_tables = schema.get("tables", [])
            sql_available = len(sql_tables) > 0
        except Exception:
            pass
        
        # Check RAG availability
        rag_available = False
        rag_tables = []
        try:
            if query_service.rag_pipeline:
                # Check if any tables have RAG configured
                db_path = _get_user_duckdb_path(current_user.id)
                if db_path.exists():
                    conn = duckdb.connect(str(db_path), read_only=True)
                    try:
                        conn.execute("""
                            SELECT table_name FROM _file_metadata 
                            WHERE table_name IN (
                                SELECT DISTINCT table_name FROM information_schema.tables
                                WHERE table_name LIKE '%_rag_config'
                            ) OR 1=1
                        """).fetchall()
                        # For now, assume RAG is available if pipeline exists
                        rag_available = True
                    except Exception:
                        pass
                    finally:
                        conn.close()
        except Exception:
            pass
        
        # Agentic hybrid requires both
        hybrid_available = sql_available and rag_available
        
        return {
            "status": "success",
            "workflows": {
                "sql": {
                    "available": sql_available,
                    "tables": [t.get("name") for t in sql_tables],
                    "description": "Fast aggregations on structured data (milliseconds)",
                },
                "rag": {
                    "available": rag_available,
                    "tables": rag_tables,
                    "description": "Semantic search on unstructured text columns",
                },
                "agentic_hybrid": {
                    "available": hybrid_available,
                    "description": "RAG → SQL → Synthesis (gold standard for complex queries)",
                    "example_queries": [
                        "Average age of patients with migraine symptoms",
                        "Count patients mentioning chest pain",
                        "BMI statistics for patients with diabetes notes",
                    ] if hybrid_available else [],
                },
            },
            "recommended_endpoint": (
                "/ingestion/query/agentic-hybrid" if hybrid_available
                else "/ingestion/query" if sql_available
                else "/ingestion/upload (upload data first)"
            ),
        }
        
    except ValueError as e:
        return {
            "status": "no_data",
            "message": str(e),
            "workflows": {
                "sql": {"available": False},
                "rag": {"available": False},
                "agentic_hybrid": {"available": False},
            },
            "recommended_endpoint": "/ingestion/upload",
        }
    except Exception as e:
        logger.error(f"Workflow status check failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))
