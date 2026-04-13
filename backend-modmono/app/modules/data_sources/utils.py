"""
Utilities for data source processing.

DuckDB file handling, schema normalization, and helpers for large file processing.
"""
import os
import re
import csv
import json
import shutil
import logging
import unicodedata
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import duckdb
import threading

# Global lock for DuckDB write operations to prevent concurrent access issues
_duckdb_write_lock = threading.Lock()

from app.core.config import get_settings

logger = logging.getLogger(__name__)

# ==========================================
# Schema Normalization
# ==========================================

KNOWN_ABBREVIATIONS = {
    'bmi': 'bmi', 'bp': 'bp', 'hr': 'hr', 'id': 'id',
    'dob': 'dob', 'ssn': 'ssn', 'mrn': 'mrn', 'icd': 'icd',
    'cpt': 'cpt', 'npi': 'npi', 'ehr': 'ehr', 'emr': 'emr',
    'hba1c': 'hba1c', 'ldl': 'ldl', 'hdl': 'hdl',
    'ast': 'ast', 'alt': 'alt', 'wbc': 'wbc', 'rbc': 'rbc',
}

MAX_COLUMN_LENGTH = 63


def normalize_column_name(col: str, index: int = 0) -> str:
    """
    Normalize a column name to a SQL-safe identifier.
    
    Transformations:
    1. Strip whitespace
    2. Convert to lowercase
    3. Replace spaces and special chars with underscores
    4. Remove parentheses content or convert to suffix
    5. Collapse multiple underscores
    6. Ensure doesn't start with number
    7. Truncate long names
    """
    if not col or not col.strip():
        return f"col_{index}"
    
    # Normalize unicode
    name = unicodedata.normalize('NFKD', col)
    name = name.encode('ASCII', 'ignore').decode('ASCII')
    
    # Lowercase and strip
    name = name.lower().strip()
    
    # Handle parentheses - convert to suffix
    name = re.sub(r'\(([^)]+)\)', r'_\1', name)
    
    # Replace non-alphanumeric with underscore
    name = re.sub(r'[^a-z0-9_]', '_', name)
    
    # Collapse multiple underscores
    name = re.sub(r'_+', '_', name)
    
    # Strip leading/trailing underscores
    name = name.strip('_')
    
    # Ensure doesn't start with number
    if name and name[0].isdigit():
        name = f"col_{name}"
    
    # Handle empty result
    if not name:
        name = f"col_{index}"
    
    # Truncate if too long
    if len(name) > MAX_COLUMN_LENGTH:
        name = name[:MAX_COLUMN_LENGTH]
    
    return name


def normalize_table_name(filename: str) -> str:
    """Convert filename to valid SQL table name."""
    # Remove extension
    name = os.path.splitext(filename)[0]
    return normalize_column_name(name)


# ==========================================
# Path Helpers
# ==========================================

def get_agent_data_dir(agent_id: str) -> Path:
    """Get the directory for an agent's data files."""
    settings = get_settings()
    agent_dir = settings.duckdb_path / f"agent_{agent_id}"
    agent_dir.mkdir(parents=True, exist_ok=True)
    return agent_dir


def get_agent_duckdb_path(agent_id: str) -> Path:
    """Get path to an agent's DuckDB file."""
    return get_agent_data_dir(agent_id) / "database.duckdb"


def get_agent_csv_path(agent_id: str, table_name: str) -> Path:
    """Get path where an agent's CSV file will be stored."""
    return get_agent_data_dir(agent_id) / f"{table_name}.csv"


# Legacy user-based functions (backward compatibility)
def get_user_data_dir(user_id: str) -> Path:
    """Get directory for a user's data files."""
    settings = get_settings()
    user_dir = settings.duckdb_path / f"user_{user_id}"
    user_dir.mkdir(parents=True, exist_ok=True)
    return user_dir


def get_user_duckdb_path(user_id: str) -> Path:
    """Get path to a user's DuckDB file."""
    return get_user_data_dir(user_id) / "database.duckdb"


def get_user_csv_path(user_id: str, table_name: str) -> Path:
    """Get path where a user's CSV file will be stored."""
    return get_user_data_dir(user_id) / f"{table_name}.csv"


# ==========================================
# Excel to CSV Streaming
# ==========================================

def stream_excel_to_csv(
    xlsx_path: str,
    csv_path: str,
    chunk_log_interval: int = 100000
) -> Dict[str, Any]:
    """
    Stream Excel file to CSV using openpyxl read-only mode.
    Avoids loading entire Excel file into RAM.
    
    Returns:
        Dict with columns, row_count, elapsed_seconds
    """
    from openpyxl import load_workbook
    
    logger.info(f"Starting Excel → CSV streaming: {xlsx_path}")
    start_time = datetime.now()
    
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    
    headers: List[str] = []
    row_count = 0
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [
                    normalize_column_name(str(cell) if cell else f"col_{i}", i)
                    for i, cell in enumerate(row)
                ]
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
            else:
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = '' if cell is None else str(cell)
                writer.writerow(row_dict)
                row_count += 1
                
                if row_count % chunk_log_interval == 0:
                    elapsed = (datetime.now() - start_time).total_seconds()
                    rate = row_count / elapsed if elapsed > 0 else 0
                    logger.info(f"  Processed {row_count:,} rows ({rate:,.0f} rows/sec)")
    
    wb.close()
    
    elapsed = (datetime.now() - start_time).total_seconds()
    logger.info(f"Excel → CSV complete: {row_count:,} rows in {elapsed:.1f}s")
    
    return {
        "columns": headers,
        "row_count": row_count,
        "elapsed_seconds": elapsed,
    }


# ==========================================
# DuckDB Operations
# ==========================================

def register_csv_in_duckdb(
    user_id: str,
    table_name: str,
    csv_path: str,
    original_filename: str,
    columns: List[str],
    row_count: int,
) -> None:
    """
    Register a CSV file in DuckDB as a virtual table.
    DuckDB queries CSV directly from disk without loading into RAM.
    
    Uses a lock to prevent concurrent connection conflicts.
    """
    db_path = get_user_duckdb_path(user_id)
    
    # Use lock to prevent "different configuration" errors when multiple
    # connections try to access the same database file
    with _duckdb_write_lock:
        conn = duckdb.connect(str(db_path), read_only=False)
        
        try:
            # Create metadata table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS _file_metadata (
                    table_name VARCHAR PRIMARY KEY,
                    original_filename VARCHAR,
                    file_type VARCHAR,
                    csv_path VARCHAR,
                    row_count BIGINT,
                    columns JSON,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Remove old entry
            conn.execute("DELETE FROM _file_metadata WHERE table_name = ?", [table_name])
            
            # Drop old view
            conn.execute(f"DROP VIEW IF EXISTS {table_name}")
            
            # Create VIEW that reads directly from CSV (virtualized)
            csv_path_escaped = str(csv_path).replace("'", "''")
            conn.execute(f"""
                CREATE VIEW {table_name} AS 
                SELECT * FROM read_csv_auto('{csv_path_escaped}', header=true)
            """)
            
            # Store metadata
            conn.execute("""
                INSERT INTO _file_metadata (table_name, original_filename, file_type, csv_path, row_count, columns)
                VALUES (?, ?, ?, ?, ?, ?)
            """, [table_name, original_filename, 'csv', str(csv_path), row_count, json.dumps(columns)])
            
            logger.info(f"Registered CSV as DuckDB view: {table_name} ({row_count:,} rows)")
        
        finally:
            conn.close()


def process_file_for_duckdb(
    user_id: str,
    table_name: str,
    source_path: str,
    file_type: str,
    original_filename: str,
) -> Dict[str, Any]:
    """
    Process a file (CSV/Excel) and register in DuckDB.
    
    Returns:
        Dict with columns, row_count, csv_path, duckdb_path
    """
    csv_path = get_user_csv_path(user_id, table_name)
    
    if file_type == 'xlsx':
        result = stream_excel_to_csv(source_path, str(csv_path))
        columns = result["columns"]
        row_count = result["row_count"]
    elif file_type == 'csv':
        shutil.copy(source_path, csv_path)
        # Get row count and columns using DuckDB
        conn = duckdb.connect(":memory:")
        csv_path_escaped = str(csv_path).replace("'", "''")
        info = conn.execute(f"SELECT COUNT(*) FROM read_csv_auto('{csv_path_escaped}')").fetchone()
        row_count = info[0]
        cols = conn.execute(f"DESCRIBE SELECT * FROM read_csv_auto('{csv_path_escaped}')").fetchall()
        columns = [normalize_column_name(c[0], i) for i, c in enumerate(cols)]
        conn.close()
    else:
        raise ValueError(f"Unsupported file type for DuckDB: {file_type}")
    
    # Register in DuckDB
    register_csv_in_duckdb(
        user_id=user_id,
        table_name=table_name,
        csv_path=str(csv_path),
        original_filename=original_filename,
        columns=columns,
        row_count=row_count,
    )
    
    return {
        "columns": columns,
        "row_count": row_count,
        "csv_path": str(csv_path),
        "duckdb_path": str(get_user_duckdb_path(user_id)),
    }


def get_file_row_count_estimate(file_path: str, file_type: str) -> int:
    """Quickly estimate row count using file size heuristics."""
    file_size = os.path.getsize(file_path)
    
    if file_type == 'csv':
        return file_size // 100  # ~100 bytes per row
    elif file_type == 'xlsx':
        return file_size // 50   # Excel compressed
    return 0


def execute_duckdb_query(
    user_id: str,
    query: str,
    max_rows: int = 10000,
) -> Dict[str, Any]:
    """
    Execute a SQL query against user's DuckDB.
    
    Returns:
        Dict with status, columns, rows, row_count, execution_time_ms
    """
    import time
    
    db_path = get_user_duckdb_path(user_id)
    
    if not db_path.exists():
        return {
            "status": "error",
            "error": "No files uploaded. Upload a CSV or Excel file first.",
            "columns": [],
            "rows": [],
            "row_count": 0,
        }
    
    try:
        start_time = time.time()
        
        conn = duckdb.connect(str(db_path), read_only=True)
        result = conn.execute(query)
        
        columns = [desc[0] for desc in result.description]
        rows_data = result.fetchmany(max_rows)
        has_more = len(rows_data) == max_rows
        
        conn.close()
        
        execution_time_ms = (time.time() - start_time) * 1000
        
        # Convert to list of dicts with JSON-safe values
        rows = []
        for row in rows_data:
            row_dict = {}
            for i, col in enumerate(columns):
                val = row[i]
                if val is None:
                    row_dict[col] = None
                elif isinstance(val, (int, float, str, bool)):
                    row_dict[col] = val
                else:
                    row_dict[col] = str(val)
            rows.append(row_dict)
        
        return {
            "status": "success",
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "execution_time_ms": round(execution_time_ms, 2),
            "error": f"Results limited to {max_rows} rows" if has_more else None,
        }
        
    except duckdb.Error as e:
        return {
            "status": "error",
            "error": f"SQL Error: {str(e)}",
            "columns": [],
            "rows": [],
            "row_count": 0,
        }
    except Exception as e:
        return {
            "status": "error",
            "error": f"Query failed: {str(e)}",
            "columns": [],
            "rows": [],
            "row_count": 0,
        }


def list_duckdb_tables(user_id: str) -> List[Dict[str, Any]]:
    """List all tables in user's DuckDB."""
    db_path = get_user_duckdb_path(user_id)
    
    if not db_path.exists():
        return []
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        
        tables_exist = conn.execute("""
            SELECT COUNT(*) FROM information_schema.tables 
            WHERE table_name = '_file_metadata'
        """).fetchone()[0]
        
        if not tables_exist:
            conn.close()
            return []
        
        rows = conn.execute("""
            SELECT table_name, original_filename, file_type, row_count, columns, created_at
            FROM _file_metadata
            ORDER BY created_at DESC
        """).fetchall()
        
        conn.close()
        
        return [
            {
                "name": row[0],
                "original_filename": row[1],
                "file_type": row[2],
                "row_count": row[3],
                "columns": json.loads(row[4]) if row[4] else [],
                "created_at": str(row[5]) if row[5] else None,
            }
            for row in rows
        ]
        
    except Exception as e:
        logger.error(f"Failed to list tables: {e}")
        return []


def get_table_schema(user_id: str, table_name: str) -> Optional[List[Dict[str, str]]]:
    """Get schema (columns and types) for a table."""
    db_path = get_user_duckdb_path(user_id)
    
    if not db_path.exists():
        return None
    
    try:
        conn = duckdb.connect(str(db_path), read_only=True)
        result = conn.execute(f"DESCRIBE SELECT * FROM {table_name}").fetchall()
        conn.close()
        
        return [
            {"column_name": row[0], "data_type": row[1]}
            for row in result
        ]
    except Exception as e:
        logger.error(f"Failed to get schema for {table_name}: {e}")
        return None


def delete_duckdb_table(user_id: str, table_name: str) -> bool:
    """Delete a table from user's DuckDB and its CSV file."""
    db_path = get_user_duckdb_path(user_id)
    
    if not db_path.exists():
        return False
    
    try:
        with _duckdb_write_lock:
            conn = duckdb.connect(str(db_path), read_only=False)
        
        csv_info = conn.execute(
            "SELECT csv_path FROM _file_metadata WHERE table_name = ?",
            [table_name]
        ).fetchone()
        
        if not csv_info:
            conn.close()
            return False
        
        csv_path = csv_info[0]
        
        conn.execute(f"DROP VIEW IF EXISTS {table_name}")
        conn.execute("DELETE FROM _file_metadata WHERE table_name = ?", [table_name])
        conn.close()
        
        if csv_path and os.path.exists(csv_path):
            os.remove(csv_path)
        
        logger.info(f"Table deleted: {table_name}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to delete table: {e}")
        return False


def delete_all_user_tables(user_id: str) -> bool:
    """Delete all tables and data for a user."""
    user_dir = get_user_data_dir(user_id)
    
    if not user_dir.exists():
        return True
    
    try:
        shutil.rmtree(user_dir, ignore_errors=True)
        logger.info(f"All tables deleted for user {user_id}")
        return True
    except Exception as e:
        logger.error(f"Failed to delete user data: {e}")
        return False


# ==========================================
# Fast Column Extraction (for large files)
# ==========================================

def extract_file_columns_fast(file_path: str, file_type: str) -> tuple:
    """
    Quickly extract column names and types from a file without loading full data.
    
    For CSV: Uses DuckDB's read_csv_auto with sample_size to infer types fast.
    For Excel: Uses openpyxl read-only mode to read only the header row.
    
    Args:
        file_path: Path to the file
        file_type: File type ('csv' or 'xlsx')
        
    Returns:
        Tuple of (column_names: List[str], column_details: List[Dict[str, str]])
        where column_details contains dicts with 'name' and 'type' keys.
    """
    if file_type == 'csv':
        return _extract_csv_columns_fast(file_path)
    elif file_type == 'xlsx':
        return _extract_excel_columns_fast(file_path)
    else:
        return [], []


def _extract_csv_columns_fast(file_path: str) -> tuple:
    """Extract columns from CSV using DuckDB sample (very fast for large files)."""
    try:
        conn = duckdb.connect(":memory:")
        csv_path_escaped = str(file_path).replace("'", "''")
        
        # DESCRIBE with sample_size only reads a small portion - fast even for huge files
        result = conn.execute(
            f"DESCRIBE SELECT * FROM read_csv_auto('{csv_path_escaped}', header=true, sample_size=1000)"
        ).fetchall()
        conn.close()
        
        columns = []
        column_details = []
        
        for i, row in enumerate(result):
            original_name = row[0]
            normalized_name = normalize_column_name(original_name, i)
            col_type = str(row[1]).upper() if row[1] else 'VARCHAR'
            
            columns.append(normalized_name)
            column_details.append({"name": normalized_name, "type": col_type})
        
        logger.info(f"Fast-extracted {len(columns)} columns from CSV: {file_path}")
        return columns, column_details
        
    except Exception as e:
        logger.error(f"Failed to extract CSV columns: {e}")
        return [], []


def _extract_excel_columns_fast(file_path: str) -> tuple:
    """Extract columns from Excel by reading only the header row."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        columns = []
        column_details = []
        
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, cell in enumerate(row):
                original_name = str(cell) if cell else f"col_{i}"
                normalized_name = normalize_column_name(original_name, i)
                
                columns.append(normalized_name)
                column_details.append({"name": normalized_name, "type": "VARCHAR"})
            break
        
        wb.close()
        logger.info(f"Fast-extracted {len(columns)} columns from Excel: {file_path}")
        return columns, column_details
        
    except Exception as e:
        logger.error(f"Failed to extract Excel columns: {e}")
        return [], []
